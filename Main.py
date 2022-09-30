from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog

import pdfplumber as pp
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
# from openpyxl.worksheet.dimensions import ColumnDimension


class Main():

    def __init__() -> None:
        filePath = None

    def _chooseFile() -> None:
        Main.filePath = filedialog.askopenfilename(
            title="Choose PDF File!",
            filetypes=[("PDF File", "*.pdf")])
      #   file = open(Main.filePath, 'r')
      #   file.close()

    def _convertFile():
        if Main.filePath == None:
            print('No File')
            return
        else:
            # get file
            _path = Main.filePath
            _pdf = pp.open(_path)

            # set pdfplumber config
            table_settings = {'vertical_strategy': 'lines',
                              'horizontal_strategy': 'lines'}

            # extracting
            _extracted = []
            _filter = ['No', 'Planning Order', 'Defect Order',
                       'Internal Repair Order', 'Defect Order']

            for page in _pdf.pages:
                _table = page.extract_table(table_settings)
                for row in _table:
                    if _filter != row:
                        _extracted.append(row)

            # sanitize data

            def _sanitizer(data):
                _sanWhiteSpace = re.sub(r"\s+", " ", data)
                _sanOrderNumber = re.sub(r"(?=[ ]\d{9})", "\n", _sanWhiteSpace)
                _sanMultiRow = re.split('\n', _sanOrderNumber)

                if (len(_sanMultiRow) == 1):
                    return _sanWhiteSpace
                return _sanMultiRow

            # process clean data
            _cleanData = []
            _index = 0

            for i in _extracted:
                _cleanData.append([i[0]])
                _cleanData[_index].append(_sanitizer(i[1]))
                _cleanData[_index].append(_sanitizer(i[2]))
                _cleanData[_index].append(_sanitizer(i[3]))
                _cleanData[_index].append(_sanitizer(i[4]))
                _index += 1

            _cleanData.insert(0, _filter)

            # create excel
            wb = Workbook()
            ws = wb.active
            ws.title = 'IOL'

            outputName = re.sub(r".pdf", "", Main.filePath)
            outputName = f"{outputName}.xlsx"

            def _merge(list, start, inc):
                columnList = ['A', 'B', 'C', 'D', 'E']

                if len(list) > 0:
                    for col in columnList:
                        if col not in list:
                            ws.merge_cells(f"{col}{start}:{col}{start+inc-1}")

            def _insert(cell, data):
                ws[cell] = data
                if 'A' not in cell:
                    ws[cell].alignment = Alignment(vertical='center', indent=1)
                else:
                    ws[cell].alignment = Alignment(
                        vertical='center', horizontal='center')

            _inc = 0  # increment
            _currentRow = 1
            _excludedCol = []  # excluded column

            for row in range(0, len(_cleanData)):
                for col in range(0, len(_cleanData[row])):
                    _chr = chr(65 + col)  # get character for cell column
                    if type(_cleanData[row][col]) == list:
                        _excludedCol.append(_chr)
                        _inc = len(_cleanData[row][col])
                        for li in range(0, len(_cleanData[row][col])):
                            cell = f"{_chr}{_currentRow + li}"
                            data = _cleanData[row][col][li]
                            _insert(cell, data)
                    else:
                        cell = f"{_chr}{_currentRow}"
                        data = _cleanData[row][col]
                        _insert(cell, data)
                    if col == 4:
                        _merge(_excludedCol, _currentRow, _inc)
                        _excludedCol = []  # reset
                if _inc != 0:
                    _currentRow += _inc
                    _inc = 0
                else:
                    _currentRow += 1

            ws.column_dimensions['A'].width = 4
            ws.column_dimensions['B'].width = 65
            ws.column_dimensions['C'].width = 65
            ws.column_dimensions['D'].width = 65
            ws.column_dimensions['E'].width = 65

            wb.save(filename=outputName)


window = Tk()
window.geometry("200x200")
window.title('IOL to Excel')
Label(window, text="-- IOL To Excel --").place(
    relx=.5, rely=.5, anchor=CENTER)
Label(window, text="Lord Maul").place(
    relx=.5, rely=.60, anchor=CENTER)
Button(window, text="Choose File", command=Main._chooseFile).place(
    relx=.5, rely=.15, anchor=CENTER)
Button(window, text="Convert", command=Main._convertFile).place(
    relx=.5, rely=.30, anchor=CENTER)


window.mainloop()
